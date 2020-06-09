import os
import openpyxl
import datetime
import PyPDF2
from PyPDF2 import PdfFileWriter, PdfFileReader

from openpyxl import Workbook


files = [f for f in os.listdir(".") if os.path.isfile(f)]
files = list(filter(lambda f: f.lower().endswith((".pdf")), files))

for pdf in files:
    with open(pdf, "rb") as f:
        inputpdf = PdfFileReader(f)

        for i in range(inputpdf.numPages):
            output = PdfFileWriter()
            output.addPage(inputpdf.getPage(i))
            name = "c:\\temp\\PDF\\Newfile\\" + pdf[:-4]+"-Page "+str(i)+".pdf"
            with open(name, "wb") as outputStream:
                output.write(outputStream)





wbs=openpyxl.load_workbook('c:\\temp\\PDF\\filelist.xlsx')

sheet=wbs.get_sheet_by_name('Sheet1')

tday=datetime.date.today()

#get the really Max_row

k=0
i=1
for i in range(0,sheet.max_row,1):

    if sheet.cell(row=i+1, column=1).value!=None :  k=k+1

row_max_real=k


print(row_max_real)



for i in range(1,row_max_real+1,1):
    

    

    Parent_file = 'spool-Page '
    Parent_file_i=Parent_file + str(i-1)+ '.pdf'
    pdfFileObj = open("c:\\temp\\PDF\\Newfile\\" + Parent_file_i, 'rb')
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
    pageObj = pdfReader.getPage(0) 
    PageTxt=(pageObj.extractText())
    print(PageTxt)
    length=len(PageTxt)
    print((PageTxt)[length-6:])
    CustomerNO=(PageTxt)[length-6:]
    pdfFileObj.close()
    Son_file_folder=sheet.cell(row=i,column=2).value
    Son_file=sheet.cell(row=i,column=1).value


    
    commandline='copy \"c:\\temp\\PDF\\NewFile\\' +   Parent_file  + str(i-1)  + '.pdf \"'   + ' \"c:\\temp\\PDF\\' +  Son_file_folder + '\\' +  'AR_Statement_' +  Son_file+ '_' + CustomerNO + '_'  + str(tday) + '.pdf\"'
    print(commandline)
    
    os.system(commandline)

